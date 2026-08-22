from __future__ import annotations

import csv
import io

import pytest
from openpyxl import Workbook

from app.ingestion.catalog import _build_column_map, ingest_catalog


def write_csv(path, text, encoding="utf-8"):
    path.write_bytes(text.encode(encoding))
    return str(path)


@pytest.mark.parametrize("delimiter", [",", ";", "\t"])
def test_csv_delimiters_and_preamble(tmp_path, delimiter):
    path = tmp_path / "supplier.csv"
    write_csv(path, f"ACME{delimiter if delimiter != ',' else ','} INDUSTRIES\nGenerated 2026\nProduct Name{delimiter}SKU{delimiter}Voltage\nMotor{delimiter}M100{delimiter}24V\n")
    result = ingest_catalog(str(path))
    assert len(result.records) == 1
    assert result.records[0].identifiers.sku == "M100"


def test_utf8_bom_and_cp1252_csv(tmp_path):
    bom = tmp_path / "bom.csv"
    bom.write_bytes("Product Name,Brand\nMótor,Síemens\n".encode("utf-8-sig"))
    assert ingest_catalog(str(bom)).records[0].brand == "Síemens"
    cp = tmp_path / "cp.csv"
    write_csv(cp, "Product Name,Brand\nMotor,Café\n", "cp1252")
    assert ingest_catalog(str(cp)).records[0].brand == "Café"


def test_supplier_aliases_and_diagnostics(tmp_path):
    path = tmp_path / "aliases.csv"
    write_csv(path, "Item,Mfr,Item Code,Rated Volts,Unit Cost\nRelay,Siemens,3RT,24V,12.50\n")
    result = ingest_catalog(str(path))
    record = result.records[0]
    assert (record.product_name, record.brand, record.identifiers.sku) == ("Relay", "Siemens", "3RT")
    assert record.commercial.price.value == 12.5
    assert result.column_stats.mapping_details["Mfr"]["target"] == "brand"
    assert result.header_row == 1 and result.rejected_rows == 0


def test_fuzzy_high_confidence_maps_but_ambiguous_does_not():
    core, specs, unknown = _build_column_map(["Product Nam", "Rated Volts", "Cost"])
    assert core["Product Nam"] == "product_name"
    assert specs["Rated Volts"] == "voltage_rating"
    assert "Cost" in unknown


@pytest.mark.parametrize("header, expected", [("SKU", "SKU-1"), ("MPN", "MPN-1")])
def test_identity_fallbacks(tmp_path, header, expected):
    path = tmp_path / "fallback.csv"
    other = "MPN" if header == "SKU" else "SKU"
    write_csv(path, f"Product Name,{header},{other}\n,{expected},\n")
    record = ingest_catalog(str(path)).records[0]
    assert record.product_name == expected


def test_xlsx_preamble_and_non_active_sheet(tmp_path):
    path = tmp_path / "catalog.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Instructions"
    ws.append(["Read me"])
    products = wb.create_sheet("Products")
    products.append(["ACME CATALOG"])
    products.append(["Product Name", "SKU"])
    products.append(["Motor", "M1"])
    wb.save(path)
    result = ingest_catalog(str(path))
    assert result.sheet == "Products"
    assert result.header_row == 2
    assert result.records[0].identifiers.sku == "M1"


def test_xlsx_blank_separator_rows_are_ignored(tmp_path):
    path = tmp_path / "catalog.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Product Name", "SKU"])
    ws.append(["Motor A", "M1"])
    ws.append([None, None])
    ws.append(["Motor B", "M2"])
    ws.append(["", ""])
    wb.save(path)

    result = ingest_catalog(str(path))

    assert result.total_rows == 2
    assert result.rejected_rows == 0
    assert result.row_warnings == []
    assert [record.product_name for record in result.records] == ["Motor A", "Motor B"]


def test_xlsm_is_accepted(tmp_path):
    path = tmp_path / "catalog.xlsm"
    wb = Workbook()
    wb.active.append(["Product Name"])
    wb.active.append(["Motor"])
    wb.save(path)
    assert ingest_catalog(str(path)).records[0].product_name == "Motor"


def test_malformed_catalog_has_clear_failure(tmp_path):
    path = tmp_path / "bad.csv"
    write_csv(path, "just metadata\nno usable columns\n")
    result = ingest_catalog(str(path))
    assert result.records == []
    assert result.rejected_rows == 1
    assert result.row_warnings
