"""Spreadsheet catalog ingestion — CSV/XLSX straight to `ProductRecord`.

Unlike PDF/DOCX/website ingestion, a catalog row is already structured: a
column header names an attribute and a cell holds its value. There is nothing
for an LLM to extract, so this module skips `IngestedDocument` and Phase 3
entirely and maps rows straight onto `ProductRecord` — the thing that makes
CSV batch scale to hundreds of rows with zero LLM calls.

Column mapping has two tiers:
- Core columns (product name, brand, sku/gtin/mpn, price, currency,
  availability, category, description) map onto `ProductRecord`'s top-level
  fields via `_CORE_COLUMNS` below.
- Everything else is resolved through `resolve_attribute()` (the same
  dictionary Phase 3/4 use) into a `Specification` entry. A column that
  resolves to neither is reported as unmapped rather than dropped silently.

Every mapped value carries provenance: `SpecSource(type=DOCUMENT,
reference=<Source.id>, snippet="row N, column 'X'")`, so Phase 5 can cite
"catalog.csv, row 5, column 'Voltage'" the same way it cites a PDF page.
"""

from __future__ import annotations

import csv
import os
import re
from difflib import SequenceMatcher
from typing import Optional

from openpyxl import load_workbook
from pydantic import BaseModel, Field

from app.schemas.attributes import ATTRIBUTE_DICTIONARY, resolve_attribute
from app.schemas.product import (
    Category,
    Commercial,
    Identifiers,
    Price,
    Provenance,
    ProductRecord,
    Source,
    SourceType,
    SpecSource,
    Specification,
    SpecStatus,
)


# Header (lower-cased, spaces/underscores folded to one form) -> ProductRecord
# top-level field. Checked before `resolve_attribute()`, so a header like
# "Brand" lands on `ProductRecord.brand` rather than becoming a spec.
_CORE_COLUMNS = {
    "product_name": "product_name",
    "name": "product_name",
    "product name": "product_name",
    "item": "product_name",
    "item name": "product_name",
    "item description": "product_name",
    "product description": "product_name",
    "title": "product_name",
    "brand": "brand",
    "manufacturer": "brand",
    "mfr": "brand",
    "make": "brand",
    "description": "description",
    "category": "category",
    "sku": "sku",
    "item code": "sku",
    "product code": "sku",
    "part code": "sku",
    "model number": "mpn",
    "gtin": "gtin",
    "upc": "gtin",
    "ean": "gtin",
    "mpn": "mpn",
    "part number": "mpn",
    "part_number": "mpn",
    "price": "price",
    "unit cost": "price",
    "currency": "currency",
    "availability": "availability",
    "stock status": "availability",
    "stock_status": "availability",
}


class ColumnMappingStats(BaseModel):
    """How the header row was understood — surfaced so a user can fix a typo'd
    column rather than silently losing data."""

    total_columns: int = 0
    mapped_core: list[str] = Field(default_factory=list)
    mapped_spec: dict[str, str] = Field(default_factory=dict)  # header -> canonical attribute
    unmapped: list[str] = Field(default_factory=list)
    mapping_details: dict[str, dict] = Field(default_factory=dict)


class CatalogIngestResult(BaseModel):
    """Output of `ingest_catalog()` — the records plus enough metadata to show
    a "500 rows, 480 mapped cleanly, 20 flagged" summary in the UI."""

    source_filename: str
    source_format: str  # "csv" | "xlsx"
    total_rows: int = 0
    records: list[ProductRecord] = Field(default_factory=list)
    column_stats: ColumnMappingStats = Field(default_factory=ColumnMappingStats)
    row_warnings: list[str] = Field(default_factory=list)  # e.g. "row 12: no product name"
    sheet: Optional[str] = None
    header_row: int = 1
    rejected_rows: int = 0
    warnings: list[str] = Field(default_factory=list)


def _normalize_header(header: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(header).strip().lower()).strip()


def _header_score(values: list[str]) -> float:
    known = 0
    nonempty = 0
    for value in values:
        if not value:
            continue
        nonempty += 1
        key = _normalize_header(value)
        if key in _CORE_COLUMNS or resolve_attribute(key):
            known += 1
    if not nonempty:
        return 0
    return known * 3 + min(nonempty, 12) * .15


def _choose_header(rows: list[list], limit: int = 25) -> tuple[int, list[str]]:
    candidates = []
    for index, raw in enumerate(rows[:limit]):
        values = ["" if v is None else str(v).strip() for v in raw]
        score = _header_score(values)
        if score and any(values[j] for j in range(len(values))):
            candidates.append((score, index, values))
    if not candidates:
        # Preserve a structurally readable header for diagnostics even when
        # none of its fields are product attributes (for example Iris data).
        readable = [
            (sum(1 for value in row if str(value).strip()), -index, index, row)
            for index, row in enumerate(rows[:limit])
            if sum(1 for value in row if str(value).strip()) >= 2
        ]
        if readable:
            _, _, index, values = max(readable)
            return index, ["" if v is None else str(v).strip() for v in values]
        return 0, []
    _, index, values = max(candidates, key=lambda item: (item[0], -item[1]))
    return index, values


def _read_csv_rows(path: str) -> tuple[list[str], list[dict[str, str]]]:
    try:
        with open(path, newline="", encoding="utf-8-sig") as f:
            sample = f.read(8192)
            f.seek(0)
            candidates = []
            for delimiter in (",", ";", "\t"):
                f.seek(0)
                parsed = list(csv.reader(f, delimiter=delimiter))
                header_index, headers = _choose_header(parsed)
                candidates.append((_header_score(headers), -header_index, parsed))
            _, _, raw_rows = max(candidates, key=lambda item: (item[0], item[1]))
    except (UnicodeDecodeError, csv.Error):
        with open(path, newline="", encoding="cp1252") as f:
            raw_rows = list(csv.reader(f, delimiter=","))
    header_index, headers = _choose_header(raw_rows)
    rows = [dict(zip(headers, row)) for row in raw_rows[header_index + 1:] if any(str(v).strip() for v in row)]
    return headers, rows


def _read_xlsx_rows(path: str) -> tuple[list[str], list[dict[str, str]], str, int]:
    wb = load_workbook(path, read_only=True, data_only=True)
    selected = None
    for sheet in wb.worksheets:
        rows = list(sheet.iter_rows(min_row=1, max_row=25, values_only=True))
        index, headers = _choose_header([list(r) for r in rows])
        score = _header_score(headers)
        if selected is None or score > selected[0]:
            selected = (score, sheet, index, headers)
    if not selected or not selected[3]:
        return [], [], wb.active.title, 1
    _, sheet, header_index, headers = selected
    rows = []
    for raw_row in sheet.iter_rows(min_row=header_index + 2, values_only=True):
        # Cells are typed by openpyxl (float/int/str/None) — stringify so
        # downstream mapping logic matches the CSV path exactly.
        row = {
            headers[i]: ("" if v is None else str(v))
            for i, v in enumerate(raw_row)
            if i < len(headers) and headers[i]
        }
        rows.append(row)
    return headers, rows, sheet.title, header_index + 1


def _build_column_map(headers: list[str]) -> tuple[dict[str, str], dict[str, str], list[str]]:
    """Classify each header once. Returns (header -> core field, header ->
    canonical attribute, unmapped headers) so per-row mapping is a dict lookup."""
    core_map: dict[str, str] = {}
    spec_map: dict[str, str] = {}
    unmapped: list[str] = []
    for header in headers:
        if not header:
            continue
        normalized = _normalize_header(header)
        core_field = _CORE_COLUMNS.get(normalized)
        if core_field:
            core_map[header] = core_field
            continue
        attr_spec = resolve_attribute(header)
        if attr_spec:
            spec_map[header] = attr_spec.attribute
        else:
            # Bounded fuzzy matching avoids turning arbitrary supplier text
            # into a field; unknown columns remain source-grounded customs.
            # Fuzzy-match only against public canonical/alias labels. Internal
            # underscore spellings are not useful candidates and can make a
            # near-exact typo lose a tie to an unresolved label.
            choices = [(key, target) for key, target in _CORE_COLUMNS.items() if "_" not in key]
            choices += [(alias, spec.attribute) for spec in ATTRIBUTE_DICTIONARY.values() for alias in spec.aliases]
            match, target = max(choices, key=lambda item: SequenceMatcher(None, normalized, _normalize_header(item[0])).ratio(), default=("", ""))
            ratio = SequenceMatcher(None, normalized, _normalize_header(match)).ratio()
            if ratio >= .9:
                if target in set(_CORE_COLUMNS.values()):
                    core_map[header] = target
                else:
                    spec_map[header] = target
            else:
                unmapped.append(header)
    return core_map, spec_map, unmapped


def _row_to_record(
    row: dict[str, str],
    row_number: int,
    core_map: dict[str, str],
    spec_map: dict[str, str], unmapped: list[str],
    source_id: str,
    filename: str,
    row_warnings: list[str],
) -> Optional[ProductRecord]:
    core_values: dict[str, str] = {}
    for header, field in core_map.items():
        value = (row.get(header) or "").strip()
        if value:
            core_values[field] = value

    product_name = core_values.get("product_name")
    if not product_name:
        # Fall back to an identifier so a row with a name typo/gap isn't
        # dropped outright — flagged via row_warnings so it's still visible.
        product_name = core_values.get("sku") or core_values.get("mpn")
    if not product_name:
        row_warnings.append(f"row {row_number}: no product name (and no sku/mpn fallback), skipped")
        return None

    price_value: Optional[float] = None
    if "price" in core_values:
        try:
            price_value = float(core_values["price"].replace(",", "").lstrip("$"))
        except ValueError:
            row_warnings.append(
                f"row {row_number}: could not parse price '{core_values['price']}'"
            )

    specifications: list[Specification] = []
    for header, attribute in spec_map.items():
        raw_value = (row.get(header) or "").strip()
        if not raw_value:
            continue
        specifications.append(
            Specification(
                attribute=attribute,
                value=raw_value,
                status=SpecStatus.EXTRACTED,
                confidence=0.95,  # direct column mapping, not model-inferred
                source=SpecSource(
                    type=SourceType.DOCUMENT,
                    reference=source_id,
                    snippet=f"row {row_number}, column '{header}'",
                ),
            )
        )
    for header in unmapped:
        raw_value = (row.get(header) or "").strip()
        if raw_value:
            specifications.append(Specification(attribute=_normalize_header(header).replace(" ", "_"), value=raw_value,
                status=SpecStatus.NEEDS_REVIEW, confidence=0.3,
                source=SpecSource(type=SourceType.DOCUMENT, reference=source_id, snippet=f"row {row_number}, column '{header}'")))

    return ProductRecord(
        product_name=product_name,
        brand=core_values.get("brand"),
        category=Category(predicted=core_values.get("category", ""), confidence=1.0 if "category" in core_values else 0.0),
        description=core_values.get("description", ""),
        identifiers=Identifiers(
            sku=core_values.get("sku"),
            gtin=core_values.get("gtin"),
            mpn=core_values.get("mpn"),
        ),
        commercial=Commercial(
            price=Price(value=price_value, currency=core_values.get("currency")) if (price_value is not None or "currency" in core_values) else None,
            availability=core_values.get("availability"),
        ),
        specifications=specifications,
        provenance=Provenance(
            sources_used=[
                Source(id=source_id, type=SourceType.DOCUMENT, reference=filename)
            ]
        ),
    )


def ingest_catalog(path: str, source_filename: str | None = None) -> CatalogIngestResult:
    """Parse a CSV or XLSX catalog file straight into `ProductRecord`s.

    No LLM call — every column is mapped deterministically via core-field
    matching or `resolve_attribute()`. Raises FileNotFoundError if `path`
    doesn't exist, ValueError for an unsupported extension.
    """
    if not os.path.isfile(path):
        raise FileNotFoundError(f"No such file: {path}")

    filename = source_filename or os.path.basename(path)
    _, ext = os.path.splitext(path)
    ext = ext.lower()

    if ext == ".csv":
        source_format = "csv"
        headers, rows = _read_csv_rows(path)
        sheet, header_row = None, 1
    elif ext in (".xlsx", ".xlsm"):
        source_format = "xlsx"
        headers, rows, sheet, header_row = _read_xlsx_rows(path)
    else:
        raise ValueError(f"Unsupported catalog file type '{ext or '(none)'}' for '{path}'. Supported: .csv, .xlsx")

    core_map, spec_map, unmapped = _build_column_map(headers)
    column_stats = ColumnMappingStats(
        total_columns=len([h for h in headers if h]),
        mapped_core=sorted(set(core_map.values())),
        mapped_spec=spec_map,
        unmapped=unmapped,
        mapping_details={h: {"target": f, "method": "alias", "confidence": 1.0} for h, f in core_map.items()} | {h: {"target": a, "method": "alias", "confidence": 1.0} for h, a in spec_map.items()} | {h: {"target": _normalize_header(h).replace(" ", "_"), "method": "custom_attribute", "confidence": 0.0} for h in unmapped},
    )

    # Imported here, not at module top level: app.validation's package init
    # (validation/groundedness.py) imports app.ingestion.models, which runs
    # this package's own __init__.py -> back to this module -> app.validation
    # while it's still mid-initialization. Deferring to call time breaks
    # that cycle. (Same root cause and fix as catalog_crawler.py's
    # extract_product import.)
    from app.validation import validate_record

    source_id = "src-1"
    records: list[ProductRecord] = []
    row_warnings: list[str] = []
    for i, row in enumerate(rows):
        # Row numbers are 1-indexed against the data rows (header excluded),
        # matching how a user reading the spreadsheet in a viewer would count.
        record = _row_to_record(row, i + 1, core_map, spec_map, unmapped, source_id, filename, row_warnings)
        if record is not None:
            records.append(validate_record(record))

    return CatalogIngestResult(
        source_filename=filename,
        source_format=source_format,
        total_rows=len(rows),
        records=records,
        column_stats=column_stats,
        row_warnings=row_warnings,
        sheet=sheet,
        header_row=header_row,
        rejected_rows=len(rows) - len(records),
        warnings=row_warnings,
    )
